import sys
import os
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

import matplotlib.cm as cm
import matplotlib.colors as mcolors


# ============================================================
# INPUT / OUTPUT
# ============================================================
INPUT_FILE = sys.argv[1] if len(sys.argv) > 1 else "dubunit_file.csv"
run_tag = os.path.splitext(os.path.basename(INPUT_FILE))[0]

OUTPUT_PDF = f"{run_tag}_source_plate_map.pdf"
OUTPUT_XLSX = f"{run_tag}_source_plate_map.xlsx"


# ============================================================
# PLATE DEFINITION (Echo source plate: 2 rows x 24 cols)
# ============================================================
PLATE_ROWS = ["A", "B"]
PLATE_COLS = list(range(1, 25))

# Serpentine / interleaved order: A01, B01, A02, B02, ...
GLOBAL_WELLS = []
for c in PLATE_COLS:
    GLOBAL_WELLS.append(f"A{c:02d}")
    GLOBAL_WELLS.append(f"B{c:02d}")


# ============================================================
# LOAD DATA
# ============================================================
df = pd.read_csv(INPUT_FILE)
df.columns = df.columns.str.strip()

required = ["Source Plate Name", "Source Well", "Sequence Name", "Transfer Volume"]
missing = [c for c in required if c not in df.columns]
if missing:
    raise ValueError(f"Missing required columns: {missing}")

plates = list(df["Source Plate Name"].unique())

HAS_DEST = "Destination Plate Name" in df.columns


# ============================================================
# GLOBAL CONSTRUCT LIST + IDs (first-seen order)
# ============================================================
ALL_CONSTRUCTS = []
for g in df["Sequence Name"]:
    if g not in ALL_CONSTRUCTS:
        ALL_CONSTRUCTS.append(g)

CONSTRUCT_ID = {name: f"C{i+1:02d}" for i, name in enumerate(ALL_CONSTRUCTS)}


# ============================================================
# PROFESSIONAL DISTINCT COLORS (tab20 + tab20b + tab20c, softened)
# ============================================================
def soften_rgba(rgba, mix_with_white=0.18, desat=0.12):
    r, g, b, a = rgba
    h, s, v = mcolors.rgb_to_hsv((r, g, b))
    s = max(0.0, s * (1.0 - desat))
    r2, g2, b2 = mcolors.hsv_to_rgb((h, s, v))
    r3 = r2 * (1.0 - mix_with_white) + 1.0 * mix_with_white
    g3 = g2 * (1.0 - mix_with_white) + 1.0 * mix_with_white
    b3 = b2 * (1.0 - mix_with_white) + 1.0 * mix_with_white
    return (r3, g3, b3, a)

def build_palette():
    palettes = []
    for name in ["tab20", "tab20b", "tab20c"]:
        cmap = cm.get_cmap(name)
        palettes.extend([cmap(i) for i in range(cmap.N)])
    palettes = [soften_rgba(rgba) for rgba in palettes]  # ~60 softened colors
    return palettes

PALETTE = build_palette()
CONSTRUCT_COLOR = {name: PALETTE[i % len(PALETTE)] for i, name in enumerate(ALL_CONSTRUCTS)}


# ============================================================
# HELPERS
# ============================================================
def is_long_gene_dest_plate(frame: pd.DataFrame) -> bool:
    """True if this plate dataframe includes any Destination Plate Name containing Long_Gene_Destination_Plate."""
    if "Destination Plate Name" not in frame.columns:
        return False
    s = frame["Destination Plate Name"].astype(str)
    return s.str.contains("Long_Gene_Destination_Plate", case=False, na=False).any()

def ordered_unique_source_wells(frame: pd.DataFrame, construct: str) -> list:
    """
    Return source wells for a construct in first-seen order.
    For Long_Gene_Destination_Plate style pooling, dedupe wells so shared subunits aren't double-counted.
    """
    sub = frame[frame["Sequence Name"] == construct]

    if is_long_gene_dest_plate(frame):
        # ✅ Fix: count each source well only once for that construct on this plate
        return sub["Source Well"].astype(str).drop_duplicates(keep="first").tolist()
    else:
        # Default behavior: each row is a unique subunit source well in practice,
        # but keep as-is in case you ever intentionally split volumes across multiple transfers.
        return sub["Source Well"].astype(str).tolist()

def first_volume_for_well(frame: pd.DataFrame, construct: str, source_well: str):
    """Pick the first transfer volume for (construct, source_well) on this plate."""
    sub = frame[(frame["Sequence Name"] == construct) & (frame["Source Well"].astype(str) == str(source_well))]
    if len(sub) == 0:
        return ""
    return sub["Transfer Volume"].iloc[0]


# ============================================================
# EXCEL SETUP
# ============================================================
wb = Workbook()
wb.remove(wb.active)
thin = Side(style="thin")
legend_rows = []


# ============================================================
# PDF GENERATION
# ============================================================
with PdfPages(OUTPUT_PDF) as pdf:

    # ✅ running counter across plates for continuity
    construct_offset = {name: 0 for name in ALL_CONSTRUCTS}

    for plate in plates:
        plate_df = df[df["Source Plate Name"] == plate].copy()

        # construct order per plate (first appearance)
        construct_order = []
        for g in plate_df["Sequence Name"]:
            if g not in construct_order:
                construct_order.append(g)

        # Build per-construct well lists (with Long_Gene dedupe fix)
        construct_wells = {c: ordered_unique_source_wells(plate_df, c) for c in construct_order}
        construct_counts = {c: len(construct_wells[c]) for c in construct_order}

        # Allocate positions on the 2x24 source plate in serpentine GLOBAL_WELLS
        assignments = {}     # plate-position well (A01/B01/...) -> construct
        mer_index = {}       # plate-position well -> global subunit index (continuous)
        well_volume = {}     # plate-position well -> volume (from source well rows)
        construct_blocks = {}  # construct -> list of plate-position wells
        ptr = 0

        for construct in construct_order:
            n = construct_counts[construct]
            positions = GLOBAL_WELLS[ptr:ptr + n]   # positions on the physical source plate
            construct_blocks[construct] = positions

            start_idx = construct_offset[construct]

            src_wells = construct_wells[construct]  # source wells from file (may include duplicates if not long-gene)

            for i, pos in enumerate(positions, start=1):
                assignments[pos] = construct
                mer_index[pos] = start_idx + i  # ✅ continuous numbering
                src_w = src_wells[i - 1]
                well_volume[pos] = first_volume_for_well(plate_df, construct, src_w)

            construct_offset[construct] += n
            ptr += n

            # Legend row (per plate)
            vols = plate_df[plate_df["Sequence Name"] == construct]["Transfer Volume"].unique()
            vol_label = vols[0] if len(vols) == 1 else "varies"
            legend_rows.append([plate, CONSTRUCT_ID[construct], construct, n, vol_label])

        # ---------------- FIGURE ----------------
        fig, ax = plt.subplots(figsize=(28, 3.3))

        # ---------------- PLATE GRID ----------------
        for r, row in enumerate(PLATE_ROWS):
            for c, col in enumerate(PLATE_COLS):
                pos = f"{row}{col:02d}"
                y = 1 - r

                if pos in assignments:
                    ax.add_patch(Rectangle(
                        (c, y), 1, 1,
                        facecolor=CONSTRUCT_COLOR[assignments[pos]],
                        edgecolor="black",
                        linewidth=0.6
                    ))
                else:
                    ax.add_patch(Rectangle(
                        (c, y), 1, 1,
                        fill=False,
                        edgecolor="black",
                        linewidth=0.6
                    ))

                if pos in mer_index:
                    ax.text(
                        c + 0.5, y + 0.38,
                        f"#{mer_index[pos]}",
                        fontsize=9.5,
                        fontweight="bold",
                        ha="center", va="center"
                    )
                    ax.text(
                        c + 0.5, y + 0.18,
                        f"{well_volume.get(pos, '')} µL",
                        fontsize=8.5,
                        ha="center", va="center"
                    )

        # ---------------- COLUMN HEADERS ----------------
        for c, col in enumerate(PLATE_COLS):
            ax.add_patch(Rectangle((c, 2), 1, 0.28, fill=False, linewidth=0.8))
            ax.text(
                c + 0.5, 2.14, str(col),
                ha="center", va="center",
                fontsize=9, fontweight="bold"
            )

        # ---------------- ROW HEADERS ----------------
        for r, row in enumerate(PLATE_ROWS):
            y = 1 - r
            ax.add_patch(Rectangle((-0.7, y), 0.7, 1, fill=False, linewidth=0.8))
            ax.text(
                -0.35, y + 0.5, row,
                ha="center", va="center",
                fontsize=11, fontweight="bold"
            )

        # ---------------- BLOCK LABELS (Construct IDs) ----------------
        for construct, positions in construct_blocks.items():
            idxs = [GLOBAL_WELLS.index(p) for p in positions]
            mid = (min(idxs) + max(idxs)) / 2
            col_mid = int(mid // 2)
            row_mid = 1 if int(mid) % 2 == 0 else 0

            ax.text(
                col_mid + 0.5, row_mid + 0.82,
                CONSTRUCT_ID[construct],
                fontsize=12,
                fontweight="bold",
                ha="center"
            )

        # ---------------- LEGEND ----------------
        legend_x0 = len(PLATE_COLS) + 0.6
        legend_y_top = 2.25
        row_step = 0.30
        constructs_per_col = 8
        box_size = 0.16
        col_width = 4.6

        for i, construct in enumerate(construct_order):
            col_idx = i // constructs_per_col
            row_idx = i % constructs_per_col
            lx = legend_x0 + col_idx * col_width
            ly = legend_y_top - row_idx * row_step

            ax.add_patch(Rectangle(
                (lx, ly - box_size),
                box_size, box_size,
                facecolor=CONSTRUCT_COLOR[construct],
                edgecolor="black",
                linewidth=0.5
            ))
            ax.text(
                lx + box_size + 0.12,
                ly - box_size / 2,
                f"{CONSTRUCT_ID[construct]}  {construct}  ({construct_counts[construct]}×100mers)",
                fontsize=9.0,
                ha="left",
                va="center"
            )

        # ---------------- TITLE ----------------
        ax.text(
            len(PLATE_COLS) / 2,
            2.48,
            f"Source Plate Map – {plate}",
            ha="center",
            va="bottom",
            fontsize=13,
            fontweight="bold"
        )

        ax.set_xlim(-0.9, len(PLATE_COLS) + 10)
        ax.set_ylim(0, 2.6)
        ax.axis("off")

        fig.tight_layout(pad=0.3)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)


# ============================================================
# EXCEL OUTPUT
# Sheet per plate = horizontal color map (IDs only)
# Legend sheet = ID, construct name, counts, volume
# ============================================================
for plate in plates:
    ws = wb.create_sheet(title=plate[:31])

    ws.append([""] + [str(c) for c in PLATE_COLS])
    for row in PLATE_ROWS:
        ws.append([row] + [""] * len(PLATE_COLS))

    plate_df = df[df["Source Plate Name"] == plate].copy()

    construct_order = []
    for g in plate_df["Sequence Name"]:
        if g not in construct_order:
            construct_order.append(g)

    construct_wells = {c: ordered_unique_source_wells(plate_df, c) for c in construct_order}
    construct_counts = {c: len(construct_wells[c]) for c in construct_order}

    ptr = 0
    for construct in construct_order:
        n = construct_counts[construct]
        positions = GLOBAL_WELLS[ptr:ptr + n]

        rgba = CONSTRUCT_COLOR[construct]
        hex_color = "%02X%02X%02X" % tuple(int(255 * x) for x in rgba[:3])
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        for pos in positions:
            r = PLATE_ROWS.index(pos[0]) + 2
            c = int(pos[1:]) + 1
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.value = CONSTRUCT_ID[construct]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ptr += n

ws_leg = wb.create_sheet(title="Legend")
ws_leg.append(["Source Plate", "Construct ID", "Construct Name", "#unique 100mers", "Volume (µL)"])
for row in legend_rows:
    ws_leg.append(row)

wb.save(OUTPUT_XLSX)

print("✅ Outputs generated:")
print(f" - {OUTPUT_PDF}")
print(f" - {OUTPUT_XLSX}")
